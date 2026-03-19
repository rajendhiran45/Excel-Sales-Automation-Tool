from __future__ import annotations

import argparse
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


REQUIRED_COLUMNS = ["Date", "Product", "Category", "Quantity", "Price", "Salesperson"]


@dataclass
class SalesReportData:
    cleaned_df: pd.DataFrame
    summary_table: pd.DataFrame
    product_sales: pd.DataFrame
    salesperson_sales: pd.DataFrame
    category_sales: pd.DataFrame


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read a sales Excel file and generate an automated sales report."
    )
    parser.add_argument("input_file", help="Path to the source sales Excel file")
    parser.add_argument(
        "--output",
        default="sales_report.xlsx",
        help="Path to the output Excel report file",
    )
    return parser.parse_args()


def load_sales_data(file_path: Path) -> pd.DataFrame:
    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    try:
        return pd.read_excel(file_path)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{file_path}': {exc}") from exc


def validate_columns(dataframe: pd.DataFrame) -> None:
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in dataframe.columns]
    if missing_columns:
        raise ValueError(
            "The input file is missing required columns: " + ", ".join(missing_columns)
        )


def clean_sales_data(dataframe: pd.DataFrame) -> pd.DataFrame:
    cleaned_df = dataframe.copy()
    cleaned_df = cleaned_df.dropna(how="all")

    cleaned_df["Date"] = pd.to_datetime(cleaned_df["Date"], errors="coerce")
    cleaned_df["Product"] = cleaned_df["Product"].fillna("Unknown").astype(str).str.strip()
    cleaned_df["Category"] = cleaned_df["Category"].fillna("Unknown").astype(str).str.strip()
    cleaned_df["Salesperson"] = (
        cleaned_df["Salesperson"].fillna("Unknown").astype(str).str.strip()
    )

    cleaned_df["Product"] = cleaned_df["Product"].replace("", "Unknown")
    cleaned_df["Category"] = cleaned_df["Category"].replace("", "Unknown")
    cleaned_df["Salesperson"] = cleaned_df["Salesperson"].replace("", "Unknown")

    cleaned_df["Quantity"] = pd.to_numeric(cleaned_df["Quantity"], errors="coerce").fillna(0)
    cleaned_df["Price"] = pd.to_numeric(cleaned_df["Price"], errors="coerce").fillna(0.0)
    cleaned_df["Total Sales"] = cleaned_df["Quantity"] * cleaned_df["Price"]

    return cleaned_df


def build_summary(
    cleaned_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    product_sales = (
        cleaned_df.groupby("Product", dropna=False)["Total Sales"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    salesperson_sales = (
        cleaned_df.groupby("Salesperson", dropna=False)["Total Sales"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    category_sales = (
        cleaned_df.groupby("Category", dropna=False)["Total Sales"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    total_revenue = float(cleaned_df["Total Sales"].sum())
    best_selling_product = product_sales.iloc[0]["Product"] if not product_sales.empty else "N/A"
    top_salesperson = (
        salesperson_sales.iloc[0]["Salesperson"] if not salesperson_sales.empty else "N/A"
    )

    summary_table = pd.DataFrame(
        [
            {"Metric": "Total Revenue", "Value": total_revenue},
            {"Metric": "Best Selling Product", "Value": best_selling_product},
            {"Metric": "Top Salesperson", "Value": top_salesperson},
        ]
    )

    return summary_table, product_sales, salesperson_sales, category_sales


def prepare_report_data(dataframe: pd.DataFrame) -> SalesReportData:
    validate_columns(dataframe)
    cleaned_df = clean_sales_data(dataframe)
    summary_table, product_sales, salesperson_sales, category_sales = build_summary(cleaned_df)
    return SalesReportData(
        cleaned_df=cleaned_df,
        summary_table=summary_table,
        product_sales=product_sales,
        salesperson_sales=salesperson_sales,
        category_sales=category_sales,
    )


def write_excel_report(
    cleaned_df: pd.DataFrame,
    summary_table: pd.DataFrame,
    product_sales: pd.DataFrame,
    salesperson_sales: pd.DataFrame,
    category_sales: pd.DataFrame,
    output_file: Path,
) -> None:
    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
            summary_table.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
            product_sales.to_excel(writer, sheet_name="Summary", index=False, startrow=6)
            salesperson_sales.to_excel(
                writer, sheet_name="Summary", index=False, startrow=6, startcol=4
            )
            category_sales.to_excel(
                writer, sheet_name="Summary", index=False, startrow=6, startcol=8
            )

        add_report_sheet(
            output_file=output_file,
            summary_table=summary_table,
            product_sales=product_sales,
            category_sales=category_sales,
        )
    except Exception as exc:
        raise RuntimeError(f"Failed to write Excel report '{output_file}': {exc}") from exc


def generate_sales_report(input_file: Path, output_file: Path) -> SalesReportData:
    sales_df = load_sales_data(input_file)
    report_data = prepare_report_data(sales_df)
    write_excel_report(
        cleaned_df=report_data.cleaned_df,
        summary_table=report_data.summary_table,
        product_sales=report_data.product_sales,
        salesperson_sales=report_data.salesperson_sales,
        category_sales=report_data.category_sales,
        output_file=output_file,
    )
    return report_data


def add_report_sheet(
    output_file: Path,
    summary_table: pd.DataFrame,
    product_sales: pd.DataFrame,
    category_sales: pd.DataFrame,
) -> None:
    workbook = load_workbook(output_file)
    if "Report" in workbook.sheetnames:
        del workbook["Report"]

    worksheet = workbook.create_sheet("Report")
    worksheet["A1"] = "Excel Sales Automation Report"
    worksheet["A1"].font = Font(bold=True, size=16)

    worksheet["A3"] = "Summary"
    worksheet["A3"].font = Font(bold=True, size=12)

    for row_index, row in enumerate(summary_table.itertuples(index=False), start=4):
        worksheet.cell(row=row_index, column=1, value=row.Metric)
        worksheet.cell(row=row_index, column=2, value=row.Value)

    worksheet["A9"] = "Sales by Category"
    worksheet["A9"].font = Font(bold=True, size=12)
    worksheet["A10"] = "Category"
    worksheet["B10"] = "Total Sales"

    for row_index, row in enumerate(category_sales.itertuples(index=False), start=11):
        worksheet.cell(row=row_index, column=1, value=row.Category)
        worksheet.cell(row=row_index, column=2, value=float(row[1]))

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        product_chart_path = temp_path / "product_sales.png"
        category_chart_path = temp_path / "category_distribution.png"

        create_product_sales_chart(product_sales, product_chart_path)
        create_category_distribution_chart(category_sales, category_chart_path)

        product_image = Image(str(product_chart_path))
        category_image = Image(str(category_chart_path))

        product_image.width = 640
        product_image.height = 360
        category_image.width = 520
        category_image.height = 360

        worksheet.add_image(product_image, "D3")
        worksheet.add_image(category_image, "D22")

        workbook.save(output_file)


def create_product_sales_chart(product_sales: pd.DataFrame, output_path: Path) -> None:
    plot_data = product_sales.copy()
    if plot_data.empty:
        plot_data = pd.DataFrame({"Product": ["No Data"], "Total Sales": [0]})

    plt.figure(figsize=(10, 5))
    plt.bar(plot_data["Product"], plot_data["Total Sales"], color="#4E79A7")
    plt.title("Product Sales")
    plt.xlabel("Product")
    plt.ylabel("Total Sales")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()


def create_category_distribution_chart(category_sales: pd.DataFrame, output_path: Path) -> None:
    plt.figure(figsize=(7, 7))
    values = category_sales["Total Sales"]
    labels = category_sales["Category"]

    if category_sales.empty or values.sum() == 0:
        values = [1]
        labels = ["No Sales Data"]

    plt.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
    plt.title("Category Distribution")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = Path(args.output)

    try:
        generate_sales_report(input_path, output_path)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Report generated successfully: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())