from pathlib import Path

from openpyxl import load_workbook

from main import generate_sales_report


INPUT_FILE = Path("sample_sales_data.xlsx")
OUTPUT_FILE = Path("sales_report_test.xlsx")
EXPECTED_SHEETS = {"Cleaned Data", "Summary", "Report"}


def run_smoke_test() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            "Missing sample_sales_data.xlsx. Run create_sample_input.py first."
        )

    generate_sales_report(INPUT_FILE, OUTPUT_FILE)

    if not OUTPUT_FILE.exists():
        raise AssertionError("Report file was not created.")

    workbook = load_workbook(OUTPUT_FILE)
    actual_sheets = set(workbook.sheetnames)
    missing_sheets = EXPECTED_SHEETS - actual_sheets
    if missing_sheets:
        raise AssertionError(f"Missing expected sheets: {sorted(missing_sheets)}")

    print("Smoke test passed.")
    print(f"Created report: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    run_smoke_test()